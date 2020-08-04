import bs4
import requests
import openpyxl


class Date:
    def __init__(self, date_soup):
        date_str = date_soup.get_text()
        href = date_soup["href"]
        split_date = date_str.split()
        months = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro",
                  "Novembro", "Dezembro"]
        self.day = int(split_date[0])
        self.month = months.index(split_date[2]) + 1
        self.year = int(split_date[4])
        self.href = href

    def __str__(self):
        return "{:02d}_{:02d}_{}".format(self.day, self.month, self.year)


class Player:
    def __init__(self, player_soup):
        player_fields_soup = player_soup.findAll("td", attrs={"data-title": True})
        self.nationality = player_fields_soup[0].find("a")["title"]
        self.overall = int(player_fields_soup[1].findAll("span")[0].get_text())
        self.potential = int(player_fields_soup[1].findAll("span")[1].get_text())
        self.name = player_fields_soup[2].find("a").get_text()
        self.age = int(player_fields_soup[4].get_text())
        self.team = player_fields_soup[6].find("a")["title"][:-8]


# year -> 19
# def scrap_page(year):
#     page_url = "https://www.fifaindex.com/pt/players/fifa{}/".format(year)
#     request_page = requests.get(page_url)
#     soup = bs4.BeautifulSoup(request_page.content, "html.parser")
#     # print(soup)
#     scrap_dates_url(soup)


def scrap_dates_url(year):
    page_url = "https://www.fifaindex.com/pt/players/fifa{}/".format(year)
    request_page = requests.get(page_url)
    soup = bs4.BeautifulSoup(request_page.content, "html.parser")

    dates = []
    months = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro",
              "Novembro", "Dezembro"]
    for a in soup.findAll("a", class_="dropdown-item"):
        for month in months:
            if month in str(a):
                dates.append(Date(a))
                break

    return dates


def scrap_premier_league(year):
    dates = scrap_dates_url(year)

    # key -> date; value -> players list
    players_date_dict = {}

    for date in dates:
        players = []
        page_index = 1
        while True:
            url = "https://www.fifaindex.com{}{}/?league=13&order=desc".format(date.href, page_index)
            print(url)
            page_players = scrap_page_players(url)
            if not page_players:
                break
            players += page_players
            page_index += 1
        players_date_dict[date] = players

    return players_date_dict


def scrap_page_players(url):
    request_page = requests.get(url)
    soup_page = bs4.BeautifulSoup(request_page.content, "html.parser")
    players_soup_list = soup_page.findAll("tr", attrs={"data-playerid": True})

    players = []

    if not players_soup_list:
        return players

    for player_soup in players_soup_list:
        players.append(Player(player_soup))

    return players


def create_wb(players, year):

    wb = openpyxl.Workbook()
    ws = wb.active
    active = True
    for date in players:
        if active:
            ws.title = str(date)
            active = False
        else:
            ws = wb.create_sheet(str(date))

        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Team")
        ws.cell(row=1, column=3, value="Overall")
        ws.cell(row=1, column=4, value="Potential")
        ws.cell(row=1, column=5, value="Nationality")
        ws.cell(row=1, column=6, value="Age")

        for player_index in range(len(players[date])):
            ws.cell(row=player_index + 2, column=1, value=players[date][player_index].name)
            ws.cell(row=player_index + 2, column=2, value=players[date][player_index].team)
            ws.cell(row=player_index + 2, column=3, value=players[date][player_index].overall)
            ws.cell(row=player_index + 2, column=4, value=players[date][player_index].potential)
            ws.cell(row=player_index + 2, column=5, value=players[date][player_index].nationality)
            ws.cell(row=player_index + 2, column=6, value=players[date][player_index].age)

    wb.save("Fifa{}PremierLeaguePlayers.xlsx".format(year))


players_fifa = scrap_premier_league(15)
create_wb(players_fifa, 15)
